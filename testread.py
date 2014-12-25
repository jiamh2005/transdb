#-*- coding:utf-8 -*    
import os
from openpyxl.reader.excel import load_workbook  
import sqlite3
import  time  

def connect_db():
    """Connects to the specific database."""
    rv = sqlite3.connect('e:/working/transdb/trans.db')
    rv.row_factory = sqlite3.Row
    return rv

def get_db():
    sqlite_db = connect_db()
    return sqlite_db

def read_excel_file(filename, db):

	#��ʼʱ��
	startTime = time.time()  

	#��ȡexcel2007�ļ�  
	wb = load_workbook(filename = r'static/Uploads/'+filename)  

	#��ʾ�ж����ű�  
	ranges = wb.get_named_ranges()  

	#ȡ��һ�ű�  
	sheetnames = wb.get_sheet_names()  
	ws = wb.get_sheet_by_name(sheetnames[0])  

	#��ʾ��������������������  
	title = ws.title  
	row = ws.get_highest_row()  
	colum = ws.get_highest_column()  

	# �����洢���ݵ��ֵ�   
	data_dic = {}   

	#�����ݴ浽�ֵ���  
	for rx in range(ws.get_highest_row()):  

		temp_list = []  
		w1 = ws.cell(row = rx+2,column = 1).value  
		w2 = ws.cell(row = rx+2,column = 2).value  
		w3 = ws.cell(row = rx+2,column = 3).value  
		temp_list = [w1,w2,w3]
		data_dic[rx] = temp_list
		if w1 == 'break':
			break

		cur = db.execute('select area, chinese, english from translations where area=? AND chinese=? AND english=?',[w1,w2,w3])
		entries = cur.fetchall()
		if not entries:
			db.execute('insert into translations (area, chinese, english) values (?, ?, ?)', [w1,w2,w3])

	db.commit()
		
	w1 = ws.cell(row = 2,column = 1).value  
	return 'Total:%d' %rx
