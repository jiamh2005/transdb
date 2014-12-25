#-*- coding:utf-8 -*    
import os
from openpyxl.reader.excel import load_workbook  
import sqlite3
import  time  

#��ʼʱ��  
startTime = time.time()  
#��ȡexcel2007�ļ�  
wb = load_workbook(filename = r'../static/Uploads/test32.xlsx')  
#��ʾ�ж����ű�  
print "Worksheet range(s):", wb.get_named_ranges()  
print "Worksheet name(s):", wb.get_sheet_names()  
#ȡ��һ�ű�  
sheetnames = wb.get_sheet_names()  
ws = wb.get_sheet_by_name(sheetnames[0])  
#��ʾ��������������������  
print "Work Sheet Titile:",ws.title  
print "Work Sheet Rows:",ws.get_highest_row()  
print "Work Sheet Cols:",ws.get_highest_column()  

# �����洢���ݵ��ֵ�   
data_dic = {}   
#�����ݴ浽�ֵ���  
for rx in range(ws.get_highest_row()):  
	  
	temp_list = []  
	pid = ws.cell(row = rx+1,column = 1).value  
	w1 = ws.cell(row = rx+1,column = 2).value  
	w2 = ws.cell(row = rx+1,column = 3).value  
	w3 = ws.cell(row = rx+1,column = 4).value  
	w4 = ws.cell(row = rx+1,column = 5).value  
	temp_list = [w1,w2,w3,w4]  
	 
	data_dic[pid] = temp_list  
#��ӡ�ֵ����ݸ���  
print 'Total:%d' %len(data_dic)
print data_dic
