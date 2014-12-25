#-*- coding:utf-8 -*  
import sqlite3 
import time  
import sys  
#workbook���  
from openpyxl.workbook import Workbook  
#����ExcelWriter���õķ�װ���˲���˵����װ�˺�ǿ���excelд�Ĺ���  
from openpyxl.writer.excel import ExcelWriter  
#һ��eggache������תΪ����ĸ�ķ���  
from openpyxl.cell import get_column_letter  
#�½�һ��workbook  
wb = Workbook()  
#�½�һ��excelWriter  
ew = ExcelWriter(workbook = wb)  
#�����ļ����·��������  
dest_filename = r'empty_book.xlsx'  
#��һ��sheet��ws  
ws = wb.worksheets[0]  
#����ws������  
ws.title = "range names"   
#¼�����ݣ�ע��col������ת��ĸ��Ȼ����Ҫ�޶�%s��string�ͣ�����������ws.cell()������ȥ,records��������Ϊһ�������ݿ����ѯ���������ݼ���  
i=1  
table = {}  
records = {'sdfasdf','xxdfd','uiui','JJJ','sOPOPdf'}
for record in records:  
    for x in range(1,len(record)+1):  
        col = get_column_letter(x)  
        ws.cell('%s%s'%(col, i)).value = '%s' % (record[x-1])        
              
    i+=1  
#�ֽ���һ��sheet��ws���ֶ�û�䣬̫ʡ�ˡ���������ȷʵ��һ���µ�sheet������Ӱ��֮ǰ�Ǹ�sheet�Ķ���  
ws = wb.create_sheet()  
ws.title = 'Pi'  
ws.cell('F5').value = 3.14  
      
#д�ļ�  
ew.save(filename = dest_filename)  